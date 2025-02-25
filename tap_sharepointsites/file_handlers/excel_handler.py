"""Handle Excel files."""

import logging
import os
import tempfile

import openpyxl

LOGGER = logging.getLogger(__name__)


class ExcelHandler:
    """Handle Excel files."""

    def __init__(self, textcontent, sheet_name, min_row, max_row, min_col, max_col):
        """Initialize ExcelHandler."""
        self.xlsheet = self._load_workbook(
            textcontent, sheet_name, min_row, max_row, min_col, max_col
        )

    def _load_workbook(
        self, textcontent, sheet_name, min_row, max_row, min_col, max_col
    ):
        """Load workbook from textcontent."""

        is_windows = os.name == 'nt'
        with tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx", delete=not is_windows) as temp:
            temp.write(textcontent)
            temp.flush()
            workbook = openpyxl.load_workbook(temp.name, read_only=True, data_only=True)
            active_sheet = workbook[sheet_name].iter_rows(
                min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
            )
            return list(active_sheet)

    def get_row_iterator(self):
        """Return a generator of rows."""
        yield from self.generator_wrapper(self.xlsheet)

    @property
    def fieldnames(self):
        """Return fieldnames."""
        fieldnames = []
        for index, cell in enumerate(self.xlsheet[0]):
            name = cell.value
            if not name:
                name = "untitled_" + str(index)
            fieldnames.append(name)
        return fieldnames

    def generator_wrapper(self, reader):
        """Wrap a reader in a generator."""
        header_row = None
        for row in reader:
            to_return = {}
            if header_row is None:
                header_row = row
                continue

            for index, cell in enumerate(row):
                header_cell = header_row[index]
                formatted_key = header_cell.value

                if not formatted_key:
                    # rename empty headers to untitled
                    formatted_key = "untitled_" + str(index)

                to_return[formatted_key] = (
                    str(cell.value) if cell.value is not None else None
                )

            yield to_return
